from copy import copy
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
import re

from openpyxl import load_workbook
from openpyxl.formatting.formatting import ConditionalFormattingList


ROOT = Path(__file__).resolve().parent
TEMPLATES = {
    "chest": ROOT / "templates" / "chest-roster.xlsx",
    "stomach": ROOT / "templates" / "stomach-roster.xlsx",
}
MAX_ROSTER_ROWS = 993  # Template detail rows 8 through 1000.
CONDITIONAL_FORMAT_RANGES = {
    "B8:H1048576": "B8:H1000",
    "F8:F1048576": "F8:F1000",
}


def japanese_date(value):
    if isinstance(value, str):
        try:
            value = date.fromisoformat(value)
        except ValueError:
            value = date.today()
    if isinstance(value, datetime):
        value = value.date()
    if not isinstance(value, date):
        value = date.today()
    return f"{value.year}年{value.month}月{value.day}日"


def safe_text(value, limit=200):
    text = str(value or "").strip()[:limit]
    if text.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


def safe_age(value):
    try:
        age = int(value)
    except (TypeError, ValueError):
        return ""
    return age if 0 <= age <= 130 else ""


def safe_film_number(value):
    text = str(value or "").strip()
    if re.fullmatch(r"\d+", text):
        number = int(text)
        if number <= 999_999_999_999_999:
            return number
    return safe_text(text, 40)


def copy_row_style(sheet, source_row, target_row, min_col=3, max_col=7):
    for col in range(min_col, max_col + 1):
        source = sheet.cell(source_row, col)
        target = sheet.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)


def limit_conditional_formatting_ranges(sheet):
    entries = [
        (CONDITIONAL_FORMAT_RANGES.get(str(item.sqref), str(item.sqref)), list(item.rules))
        for item in sheet.conditional_formatting
    ]
    sheet.conditional_formatting = ConditionalFormattingList()
    for cell_range, rules in entries:
        for rule in rules:
            sheet.conditional_formatting.add(cell_range, copy(rule))


def film_number_sort_key(item):
    value = str(item.get("filmNumber") or "").strip()
    parts = re.split(r"(\d+)", value.casefold())
    natural = tuple((0, int(part)) if part.isdigit() else (1, part) for part in parts if part)
    return (not bool(value), natural)


def dust_label(item):
    if item.get("legacyDust") or (item.get("asbestos") and "pneumoconiosis" not in item):
        return "塵肺"
    labels = []
    if item.get("pneumoconiosis"):
        labels.append("塵肺")
    if item.get("asbestos"):
        labels.append("アスベスト")
    return "・".join(labels)


def build_roster(kind, customer_name, exam_date, rows):
    if kind not in TEMPLATES:
        raise ValueError("Unknown roster type")
    if not isinstance(rows, list):
        raise ValueError("rows must be a list")
    if len(rows) > MAX_ROSTER_ROWS:
        raise ValueError(f"連名簿は{MAX_ROSTER_ROWS}件まで出力できます")

    workbook = load_workbook(TEMPLATES[kind])
    sheet = workbook.active
    rows = sorted(rows, key=film_number_sort_key)
    limit_conditional_formatting_ranges(sheet)
    sheet["C4"] = safe_text(customer_name, 100)
    sheet["E4"] = japanese_date(exam_date)

    for row in sheet.iter_rows():
        for cell in row:
            if cell.comment is not None:
                cell.comment = None
            if isinstance(cell.value, str) and cell.value.startswith("※印刷"):
                cell.value = None

    sheet.print_title_rows = "7:7"
    last_print_row = max(8, 7 + len(rows))
    sheet.print_area = f"$B$2:$G${last_print_row}"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
    if kind == "chest":
        sheet.column_dimensions["G"].width = 18

    for row_number in range(8, 1001):
        for col in range(3, 8):
            sheet.cell(row_number, col).value = None
        sheet.cell(row_number, 9).value = None

    for index, item in enumerate(rows, start=8):
        if index > 8:
            copy_row_style(sheet, 8, index)
        sheet.cell(index, 3).value = safe_text(item.get("nameKana") or item.get("name"), 120)
        sheet.cell(index, 4).value = safe_text(item.get("sex"), 20)
        sheet.cell(index, 5).value = safe_age(item.get("age"))
        film_number = safe_film_number(item.get("filmNumber"))
        sheet.cell(index, 6).value = film_number
        sheet.cell(index, 6).number_format = "0" if isinstance(film_number, int) else "@"
        if kind == "chest":
            sheet.cell(index, 7).value = dust_label(item) or None
        if index == 8:
            sheet.cell(index, 9).value = '=IF(OR(F6+1=F8,F6="",F8=""),"","★")'
        else:
            previous = index - 1
            sheet.cell(index, 9).value = f'=IF(OR(F{previous}+1=F{index},F{previous}="",F{index}=""),"","★")'

    sheet["F3"] = "=COUNTA($F$8:F1000)"
    if kind == "chest":
        sheet["F2"] = "=COUNTA($G$8:G1000)"
    if workbook.calculation:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
