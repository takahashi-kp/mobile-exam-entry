import unittest
from datetime import date

from openpyxl import load_workbook

from roster_export import MAX_ROSTER_ROWS, build_roster


SAMPLE_ROWS = [
    {"nameKana": "ヤマダ タロウ", "sex": "男性", "age": 52, "filmNumber": "00123", "asbestos": True},
    {"nameKana": "スズキ ハナコ", "sex": "女性", "age": 47, "filmNumber": "00124", "asbestos": False},
]


class RosterExportTests(unittest.TestCase):
    def test_chest_roster_populates_template_and_counts(self):
        output = build_roster("chest", "テスト顧客", date(2026, 7, 14), SAMPLE_ROWS)
        sheet = load_workbook(output, data_only=False).active

        self.assertEqual(sheet["C4"].value, "テスト顧客")
        self.assertEqual(sheet["E4"].value, "2026年7月14日")
        self.assertEqual(sheet["F2"].value, '=COUNTIF($G$8:G1000,"塵肺")')
        self.assertEqual(sheet["F3"].value, "=COUNTA($F$8:F1000)")
        self.assertEqual(sheet["C8"].value, "ヤマダ タロウ")
        self.assertEqual(sheet["F8"].value, "00123")
        self.assertEqual(sheet["F8"].number_format, "@")
        self.assertEqual(sheet["G8"].value, "塵肺")
        self.assertIsNone(sheet["G9"].value)

    def test_stomach_roster_has_no_asbestos_column_value(self):
        output = build_roster("stomach", "テスト顧客", "2026-07-14", SAMPLE_ROWS)
        sheet = load_workbook(output, data_only=False).active

        self.assertEqual(sheet.title, "胃部")
        self.assertEqual(sheet["F8"].value, "00123")
        self.assertIsNone(sheet["G8"].value)

    def test_row_limit_is_enforced(self):
        with self.assertRaises(ValueError):
            build_roster("chest", "顧客", "2026-07-14", [{}] * (MAX_ROSTER_ROWS + 1))

    def test_excel_formula_injection_is_written_as_text(self):
        output = build_roster("stomach", "=HYPERLINK(\"bad\")", "2026-07-14", [{"nameKana": "+CMD", "filmNumber": "@001"}])
        sheet = load_workbook(output, data_only=False).active

        self.assertEqual(sheet["C4"].value, "'=HYPERLINK(\"bad\")")
        self.assertEqual(sheet["C8"].value, "'+CMD")
        self.assertEqual(sheet["F8"].value, "'@001")

    def test_roster_is_sorted_and_matches_print_settings(self):
        rows = [
            {"nameKana": "サン", "filmNumber": "10"},
            {"nameKana": "イチ", "filmNumber": "1"},
            {"nameKana": "ニ", "filmNumber": "2"},
        ]

        output = build_roster("chest", "", "2026-07-14", rows)
        sheet = load_workbook(output, data_only=False).active

        self.assertEqual([sheet[f"F{row}"].value for row in range(8, 11)], ["1", "2", "10"])
        self.assertEqual(sheet.print_title_rows, "$7:$7")
        self.assertEqual(sheet.page_setup.paperSize, 9)
        self.assertEqual(sheet.page_setup.orientation, "portrait")
        self.assertFalse(any(cell.comment for row in sheet.iter_rows() for cell in row))
        self.assertFalse(any(isinstance(cell.value, str) and cell.value.startswith("※印刷") for row in sheet.iter_rows() for cell in row))


if __name__ == "__main__":
    unittest.main()
